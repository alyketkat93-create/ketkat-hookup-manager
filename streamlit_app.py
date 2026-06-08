import io
import os
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# =========================================================
# KETKAT Streamlit MVP V3
# Light UI + Logo + Export Center + BOQ to Rate Analysis
# =========================================================

APP_TITLE = "KETKAT Hook-up & Installation Manager"
DEFAULT_SHEET_NAME = "BOQ_Data"
LOGO_PATH = "assets/ketkat_logo.png"

REQUIRED_COLUMNS = [
    "Main Category",
    "Sub System",
    "Family",
    "Work Type",
    "Title",
    "No",
    "Description",
    "Unit",
    "Default Qty",
    "Unit Price",
]


st.set_page_config(
    page_title=APP_TITLE,
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)


# =========================================================
# Styling
# =========================================================
st.markdown(
    """
    <style>
    :root {
        --ketkat-navy: #0B172B;
        --ketkat-navy-2: #12213D;
        --ketkat-gold: #D9A514;
        --ketkat-light: #F6F8FB;
        --ketkat-border: #D9E2EC;
        --ketkat-muted: #64748B;
        --ketkat-text: #0F172A;
        --ketkat-success: #16A34A;
    }

    .stApp {
        background: #F6F8FB;
        color: var(--ketkat-text);
    }

    [data-testid="stSidebar"] {
        background: #0B172B;
    }

    [data-testid="stSidebar"] * {
        color: #F8FAFC !important;
    }

    [data-testid="stSidebar"] hr {
        border-color: #1D2A44;
    }

    .block-container {
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1600px;
    }

    .hero {
        background: linear-gradient(135deg, #0B172B 0%, #12213D 65%, #1E3A5F 100%);
        border-radius: 18px;
        padding: 26px 30px;
        border: 1px solid #1D2A44;
        margin-bottom: 22px;
        box-shadow: 0 12px 28px rgba(15, 23, 42, 0.12);
    }

    .hero-title {
        font-size: 34px;
        font-weight: 900;
        color: #FFFFFF;
        margin: 0;
        line-height: 1.12;
    }

    .hero-subtitle {
        font-size: 15px;
        color: #CBD5E1;
        margin-top: 8px;
    }

    .hero-badge {
        display: inline-block;
        background: #D9A514;
        color: #0B172B;
        font-size: 12px;
        font-weight: 900;
        padding: 7px 12px;
        border-radius: 999px;
        margin-bottom: 12px;
        letter-spacing: .4px;
    }

    .section-title {
        font-size: 22px;
        font-weight: 900;
        color: #0F172A;
        margin-top: 10px;
        margin-bottom: 12px;
    }

    .section-caption {
        color: #64748B;
        font-size: 14px;
        margin-top: -6px;
        margin-bottom: 16px;
    }

    .ketkat-card {
        background: #FFFFFF;
        padding: 18px 20px;
        border-radius: 16px;
        border: 1px solid #D9E2EC;
        min-height: 118px;
        box-shadow: 0 8px 22px rgba(15, 23, 42, 0.06);
    }

    .ketkat-card-label {
        color: #64748B;
        font-size: 13px;
        font-weight: 800;
        margin-bottom: 8px;
    }

    .ketkat-card-value {
        color: #0B172B;
        font-size: 25px;
        font-weight: 900;
        line-height: 1.12;
        word-break: break-word;
    }

    .ketkat-card-value.gold {
        color: #D9A514;
    }

    .ketkat-card-sub {
        color: #64748B;
        font-size: 12px;
        margin-top: 8px;
    }

    .info-box {
        background: #FFFFFF;
        border-left: 5px solid #D9A514;
        border-radius: 14px;
        padding: 16px 18px;
        border-top: 1px solid #D9E2EC;
        border-right: 1px solid #D9E2EC;
        border-bottom: 1px solid #D9E2EC;
        box-shadow: 0 8px 22px rgba(15, 23, 42, 0.05);
        margin-bottom: 16px;
    }

    .info-box-title {
        font-size: 16px;
        font-weight: 900;
        color: #0B172B;
        margin-bottom: 4px;
    }

    .info-box-text {
        font-size: 13px;
        color: #64748B;
    }

    div[data-testid="stDataFrame"],
    div[data-testid="stDataEditor"] {
        border-radius: 16px;
        overflow: hidden;
        border: 1px solid #D9E2EC;
        box-shadow: 0 8px 22px rgba(15, 23, 42, 0.05);
    }

    .stButton > button,
    .stDownloadButton > button {
        border-radius: 10px !important;
        font-weight: 800 !important;
        border: 0 !important;
        padding: 0.65rem 1.2rem !important;
    }

    .stDownloadButton > button {
        background: #D9A514 !important;
        color: #0B172B !important;
    }

    .stButton > button[kind="primary"] {
        background: #0B172B !important;
        color: #FFFFFF !important;
    }

    label, .stSelectbox label, .stTextInput label, .stNumberInput label {
        color: #0F172A !important;
        font-weight: 800 !important;
    }

    .footer-note {
        color: #64748B;
        font-size: 12px;
        text-align: center;
        margin-top: 28px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# Helpers
# =========================================================
def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))

    df = df.copy()

    for col in [
        "Main Category",
        "Sub System",
        "Family",
        "Work Type",
        "Title",
        "Description",
        "Unit",
    ]:
        df[col] = df[col].fillna("").astype(str).str.strip()

    df["No"] = pd.to_numeric(df["No"], errors="coerce").fillna(0).astype(int)
    df["Default Qty"] = pd.to_numeric(df["Default Qty"], errors="coerce").fillna(0.0)
    df["Unit Price"] = pd.to_numeric(df["Unit Price"], errors="coerce").fillna(0.0)

    df["_row_id"] = range(1, len(df) + 1)
    return df


def read_excel_file(uploaded_file) -> pd.DataFrame:
    excel = pd.ExcelFile(uploaded_file)
    sheet_name = DEFAULT_SHEET_NAME if DEFAULT_SHEET_NAME in excel.sheet_names else excel.sheet_names[0]
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
    return normalize_df(df)


def unique_options(series: pd.Series) -> list[str]:
    values = sorted([str(x).strip() for x in series.dropna().unique() if str(x).strip()])
    return ["All"] + values


def money(value: float) -> str:
    try:
        return f"SAR {float(value):,.2f}"
    except Exception:
        return "SAR 0.00"


def card(label: str, value: str, sub: str = "", gold: bool = False):
    gold_class = " gold" if gold else ""
    st.markdown(
        f"""
        <div class="ketkat-card">
            <div class="ketkat-card-label">{label}</div>
            <div class="ketkat-card-value{gold_class}">{value}</div>
            <div class="ketkat-card-sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def info_box(title: str, text: str):
    st.markdown(
        f"""
        <div class="info-box">
            <div class="info-box-title">{title}</div>
            <div class="info-box-text">{text}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def excel_format_common(ws, export_df, project_info, selection_text, title="KETKAT Hook-up & Installation Manager"):
    ws["A1"] = title
    ws["A2"] = project_info.get("project_name", "Project BOQ")
    ws["A3"] = f"Selection: {selection_text}"
    ws["A4"] = (
        f"Client: {project_info.get('client', '')} | "
        f"Consultant: {project_info.get('consultant', '')} | "
        f"Prepared By: {project_info.get('prepared_by', '')} | "
        f"Date: {datetime.now().strftime('%Y-%m-%d')}"
    )

    fill_dark = PatternFill("solid", fgColor="0B172B")
    fill_gold = PatternFill("solid", fgColor="D9A514")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ["A1", "A2", "A3", "A4"]:
        ws[cell].fill = fill_dark
        ws[cell].font = Font(color="FFFFFF", bold=True)

    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A2"].font = Font(color="FFFFFF", bold=True, size=12)

    header_row = 6
    for cell in ws[header_row]:
        cell.fill = fill_gold
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    last_row = header_row + len(export_df)
    if last_row >= header_row + 1:
        for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row, min_col=1, max_col=len(export_df.columns)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if cell.column in [1, 3, 4]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.column in [5, 6]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "A7"


def build_boq_export_excel(export_df: pd.DataFrame, project_info: dict, selection_text: str) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, startrow=5, sheet_name="BOQ")
        ws = writer.sheets["BOQ"]
        excel_format_common(ws, export_df, project_info, selection_text)

        total_row = 6 + len(export_df) + 2
        ws[f"E{total_row}"] = "Total"
        ws[f"F{total_row}"] = float(export_df["T/P (SAR)"].sum()) if not export_df.empty else 0

        fill_dark = PatternFill("solid", fgColor="0B172B")
        thin = Side(style="thin", color="D9E2EC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in [ws[f"E{total_row}"], ws[f"F{total_row}"]]:
            cell.fill = fill_dark
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="center")

        widths = {"A": 10, "B": 80, "C": 14, "D": 12, "E": 18, "F": 18}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    return output.getvalue()


def build_database_summary_excel(df: pd.DataFrame, project_info: dict) -> bytes:
    output = io.BytesIO()

    summary = (
        df.groupby(["Main Category", "Sub System", "Family", "Work Type"], dropna=False)
        .agg(
            Items=("Description", "count"),
            Default_Qty_Total=("Default Qty", "sum"),
            Unit_Price_Total=("Unit Price", "sum"),
        )
        .reset_index()
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Database Summary")
        df.drop(columns=["_row_id"], errors="ignore").to_excel(writer, index=False, sheet_name="Full Database")

        for sheet in writer.sheets.values():
            for col in sheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[col_letter].width = min(max_len + 3, 60)

            for cell in sheet[1]:
                cell.fill = PatternFill("solid", fgColor="D9A514")
                cell.font = Font(color="000000", bold=True)
                cell.alignment = Alignment(horizontal="center")

    return output.getvalue()


def build_rate_analysis_excel(resources_df: pd.DataFrame, info: dict) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        resources_df.to_excel(writer, index=False, startrow=7, sheet_name="Rate Analysis")
        ws = writer.sheets["Rate Analysis"]

        ws["A1"] = "KETKAT Rate Analysis"
        ws["A2"] = f"Analysis No: {info.get('analysis_no', '')}"
        ws["A3"] = f"Description: {info.get('description', '')}"
        ws["A4"] = f"Unit: {info.get('unit', '')}"
        ws["A5"] = f"Markup Mode: {info.get('mode', '')} | Overhead: {info.get('overhead_pct', 0)}% | Profit: {info.get('profit_pct', 0)}%"
        ws["A6"] = f"Final Rate: SAR {info.get('final_rate', 0):,.2f}"

        fill_dark = PatternFill("solid", fgColor="0B172B")
        fill_gold = PatternFill("solid", fgColor="D9A514")
        thin = Side(style="thin", color="D9E2EC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ["A1", "A2", "A3", "A4", "A5", "A6"]:
            ws[cell].fill = fill_dark
            ws[cell].font = Font(color="FFFFFF", bold=True)

        ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)

        header_row = 8
        for cell in ws[header_row]:
            cell.fill = fill_gold
            cell.font = Font(color="000000", bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + len(resources_df), min_col=1, max_col=len(resources_df.columns)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        widths = {"A": 16, "B": 16, "C": 45, "D": 12, "E": 12, "F": 16, "G": 16}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    return output.getvalue()


def build_filtered_work_df(df: pd.DataFrame, main: str, sub: str, family: str, work_type: str, search: str) -> pd.DataFrame:
    filtered = df.copy()

    if main != "All":
        filtered = filtered[filtered["Main Category"] == main]
    if sub != "All":
        filtered = filtered[filtered["Sub System"] == sub]
    if family != "All":
        filtered = filtered[filtered["Family"] == family]
    if work_type != "All":
        filtered = filtered[filtered["Work Type"] == work_type]

    if search.strip():
        keyword = search.strip().lower()
        mask = (
            filtered["Description"].astype(str).str.lower().str.contains(keyword, na=False)
            | filtered["Title"].astype(str).str.lower().str.contains(keyword, na=False)
            | filtered["Unit"].astype(str).str.lower().str.contains(keyword, na=False)
        )
        filtered = filtered[mask]

    work_df = filtered.copy()
    if not work_df.empty:
        work_df["QTY"] = work_df["Default Qty"]
        work_df["U/R (SAR)"] = work_df["Unit Price"]
        work_df["T/P (SAR)"] = work_df["QTY"] * work_df["U/R (SAR)"]

    return work_df


# =========================================================
# Session State
# =========================================================
if "database_df" not in st.session_state:
    st.session_state.database_df = pd.DataFrame()

if "loaded_filename" not in st.session_state:
    st.session_state.loaded_filename = ""

if "project_info" not in st.session_state:
    st.session_state.project_info = {
        "project_name": "Demo Project",
        "client": "",
        "consultant": "",
        "location": "",
        "prepared_by": "KETKAT Contracting",
    }

if "last_boq_df" not in st.session_state:
    st.session_state.last_boq_df = pd.DataFrame()

if "last_selection_text" not in st.session_state:
    st.session_state.last_selection_text = "All / All / All / All"

if "rate_item" not in st.session_state:
    st.session_state.rate_item = {
        "description": "",
        "unit": "Item",
        "unit_rate": 0.0,
        "no": "",
    }


# =========================================================
# Sidebar
# =========================================================
with st.sidebar:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=120)

    st.markdown("## 🏗️ KETKAT")
    st.caption("Hook-up & Installation Manager")

    page = st.radio(
        "Navigation",
        ["BOQ Builder", "Rate Analysis", "Project Info", "Database Preview", "Export Center"],
        index=0,
    )

    st.divider()

    uploaded_logo = st.file_uploader("Upload Logo Optional", type=["png", "jpg", "jpeg"], key="logo_uploader")
    if uploaded_logo:
        st.image(uploaded_logo, width=120)
        st.caption("Logo preview only. To make it permanent, save it as assets/ketkat_logo.png.")

    uploaded_file = st.file_uploader(
        "Upload BOQ Database",
        type=["xlsx", "xls"],
        help="Required columns: Main Category, Sub System, Family, Work Type, Title, No, Description, Unit, Default Qty, Unit Price",
        key="db_uploader",
    )

    if uploaded_file:
        try:
            st.session_state.database_df = read_excel_file(uploaded_file)
            st.session_state.loaded_filename = uploaded_file.name
            st.success(f"Loaded: {uploaded_file.name}")
        except Exception as exc:
            st.error(f"Failed to load database: {exc}")

    st.divider()

    if not st.session_state.database_df.empty:
        st.caption(f"Database file: {st.session_state.loaded_filename}")
        st.caption(f"Rows: {len(st.session_state.database_df):,}")
    else:
        st.warning("No database loaded yet.")

    st.divider()
    st.caption("V3 Light MVP")


# =========================================================
# Hero Header
# =========================================================
st.markdown(
    f"""
    <div class="hero">
        <div class="hero-badge">MEP ESTIMATION WEB TOOL</div>
        <div class="hero-title">{APP_TITLE}</div>
        <div class="hero-subtitle">
            Prepare Hook-up & Installation BOQ items from Excel database with smart filters,
            editable quantities/rates, Rate Analysis linkage, and professional Excel exports.
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# Pages
# =========================================================
df = st.session_state.database_df


if page == "Project Info":
    st.markdown('<div class="section-title">Project Information</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-caption">This information will appear in exported BOQ reports during the current session.</div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.session_state.project_info["project_name"] = st.text_input("Project Name", st.session_state.project_info.get("project_name", ""))
        st.session_state.project_info["client"] = st.text_input("Client", st.session_state.project_info.get("client", ""))
        st.session_state.project_info["consultant"] = st.text_input("Consultant", st.session_state.project_info.get("consultant", ""))
    with c2:
        st.session_state.project_info["location"] = st.text_input("Location", st.session_state.project_info.get("location", ""))
        st.session_state.project_info["prepared_by"] = st.text_input("Prepared By", st.session_state.project_info.get("prepared_by", ""))

    info_box("Saved for current session", "Project information is stored while the app session is active and will be used in Excel exports.")


elif page == "Database Preview":
    st.markdown('<div class="section-title">Database Preview</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-caption">Review the uploaded Excel database before building BOQ outputs.</div>', unsafe_allow_html=True)

    if df.empty:
        info_box("No database loaded", "Upload your Excel database from the sidebar to start.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            card("Total Items", f"{len(df):,}", "Rows loaded", gold=True)
        with c2:
            card("Main Categories", f"{df['Main Category'].nunique():,}", "Unique categories")
        with c3:
            card("Sub Systems", f"{df['Sub System'].nunique():,}", "Unique systems")
        with c4:
            card("Families", f"{df['Family'].nunique():,}", "Unique families")

        st.dataframe(df.drop(columns=["_row_id"], errors="ignore"), use_container_width=True, height=540)


elif page == "Export Center":
    st.markdown('<div class="section-title">Export Center</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-caption">Export the current BOQ selection, full database preview, or database summary.</div>', unsafe_allow_html=True)

    if df.empty:
        info_box("No database loaded", "Upload your Excel database from the sidebar to enable exports.")
    else:
        c1, c2, c3 = st.columns(3)
        with c1:
            card("Database Rows", f"{len(df):,}", "Loaded rows", gold=True)
        with c2:
            card("Last BOQ Rows", f"{len(st.session_state.last_boq_df):,}", "Last edited BOQ table")
        with c3:
            total_last = float(st.session_state.last_boq_df["T/P (SAR)"].sum()) if not st.session_state.last_boq_df.empty and "T/P (SAR)" in st.session_state.last_boq_df else 0.0
            card("Last BOQ Total", money(total_last), "From BOQ Builder")

        summary_bytes = build_database_summary_excel(df, st.session_state.project_info)
        st.download_button(
            "⬇️ Export Database Summary",
            data=summary_bytes,
            file_name=f"KETKAT_Database_Summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if not st.session_state.last_boq_df.empty:
            last_export_df = st.session_state.last_boq_df[["No", "Description", "Unit", "QTY", "U/R (SAR)", "T/P (SAR)"]].copy()
            last_boq_bytes = build_boq_export_excel(
                last_export_df,
                st.session_state.project_info,
                st.session_state.last_selection_text,
            )
            st.download_button(
                "⬇️ Export Last BOQ Selection",
                data=last_boq_bytes,
                file_name=f"KETKAT_Last_BOQ_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        else:
            info_box("No BOQ selection yet", "Go to BOQ Builder, filter items, edit quantities/rates, then return here to export the last selection.")


elif page == "Rate Analysis":
    st.markdown('<div class="section-title">Rate Analysis</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-caption">Build a simple unit rate analysis. You can send a selected BOQ item from BOQ Builder to this page.</div>', unsafe_allow_html=True)

    selected_item = st.session_state.rate_item

    c1, c2, c3 = st.columns(3)
    with c1:
        analysis_no = st.text_input("Analysis No", value=f"RA-{selected_item.get('no') or '001'}")
    with c2:
        unit = st.text_input("Unit", value=selected_item.get("unit", "Item") or "Item")
    with c3:
        category = st.selectbox("Category", ["Hook-up", "Installation", "Labor", "Material", "Equipment"], index=0)

    description = st.text_area("Description", value=selected_item.get("description", ""), height=90)

    base_rate = float(selected_item.get("unit_rate", 0.0) or 0.0)

    default_resources = pd.DataFrame(
        [
            {"Type": "Labor", "Code": "L-001", "Description": "Fitter / Technician", "Unit": "hr", "Qty": 1.0, "Rate": 0.0, "Amount": 0.0},
            {"Type": "Labor", "Code": "L-002", "Description": "Helper", "Unit": "hr", "Qty": 1.0, "Rate": 0.0, "Amount": 0.0},
            {"Type": "Consumable", "Code": "C-001", "Description": "Consumables", "Unit": "lot", "Qty": 1.0, "Rate": 0.0, "Amount": 0.0},
        ]
    )

    if base_rate > 0:
        default_resources.loc[0, "Rate"] = round(base_rate * 0.55, 2)
        default_resources.loc[1, "Rate"] = round(base_rate * 0.25, 2)
        default_resources.loc[2, "Rate"] = round(base_rate * 0.20, 2)
        default_resources["Amount"] = default_resources["Qty"] * default_resources["Rate"]

    st.markdown("### Resources Breakdown")
    resources_df = st.data_editor(
        default_resources,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Type": st.column_config.SelectboxColumn("Type", options=["Labor", "Material", "Equipment", "Consumable", "Subcontractor", "Other"]),
            "Qty": st.column_config.NumberColumn("Qty", min_value=0.0, step=1.0),
            "Rate": st.column_config.NumberColumn("Rate", min_value=0.0, step=10.0, format="%.2f"),
            "Amount": st.column_config.NumberColumn("Amount", format="%.2f"),
        },
    )
    resources_df["Amount"] = pd.to_numeric(resources_df["Qty"], errors="coerce").fillna(0) * pd.to_numeric(resources_df["Rate"], errors="coerce").fillna(0)

    direct_cost = float(resources_df["Amount"].sum())

    c1, c2, c3 = st.columns(3)
    with c1:
        overhead_pct = st.number_input("Overhead %", min_value=0.0, value=10.0, step=1.0)
    with c2:
        profit_pct = st.number_input("Profit %", min_value=0.0, value=15.0, step=1.0)
    with c3:
        mode = st.selectbox("Markup Mode", ["Simple", "Compound"])

    if mode == "Simple":
        final_rate = direct_cost * (1 + (overhead_pct + profit_pct) / 100)
    else:
        final_rate = direct_cost * (1 + overhead_pct / 100) * (1 + profit_pct / 100)

    c1, c2, c3 = st.columns(3)
    with c1:
        card("Direct Cost", money(direct_cost), "Resources total")
    with c2:
        card("Markup", f"{overhead_pct + profit_pct:.2f}%", f"{mode} mode")
    with c3:
        card("Final Rate", money(final_rate), "Calculated unit rate", gold=True)

    ra_bytes = build_rate_analysis_excel(
        resources_df,
        {
            "analysis_no": analysis_no,
            "description": description,
            "unit": unit,
            "mode": mode,
            "overhead_pct": overhead_pct,
            "profit_pct": profit_pct,
            "final_rate": final_rate,
        },
    )
    st.download_button(
        "⬇️ Export Rate Analysis Excel",
        data=ra_bytes,
        file_name=f"KETKAT_Rate_Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


else:
    st.markdown('<div class="section-title">BOQ Builder</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-caption">Upload your database, filter BOQ items, edit quantities/rates, send an item to Rate Analysis, and export BOQ Excel.</div>', unsafe_allow_html=True)

    if df.empty:
        info_box("Upload database to start", "Use the file uploader in the left sidebar. The Excel file must include the required BOQ database columns.")
        st.stop()

    c1, c2, c3 = st.columns(3)
    with c1:
        card("Database Items", f"{len(df):,}", "Total loaded rows", gold=True)
    with c2:
        card("Source File", st.session_state.loaded_filename or "Uploaded file", "Active database")
    with c3:
        card("Currency", "SAR", "Default pricing currency")

    st.markdown("### Smart Filters")

    f1, f2, f3, f4 = st.columns(4)

    with f1:
        main = st.selectbox("Main Category", unique_options(df["Main Category"]))

    temp = df.copy()
    if main != "All":
        temp = temp[temp["Main Category"] == main]

    with f2:
        sub = st.selectbox("Sub System", unique_options(temp["Sub System"]))

    temp2 = temp.copy()
    if sub != "All":
        temp2 = temp2[temp2["Sub System"] == sub]

    with f3:
        family = st.selectbox("Family", unique_options(temp2["Family"]))

    temp3 = temp2.copy()
    if family != "All":
        temp3 = temp3[temp3["Family"] == family]

    with f4:
        work_type = st.selectbox("Work Type", unique_options(temp3["Work Type"]))

    search = st.text_input("Search Description / Title / Unit", "")

    work_df = build_filtered_work_df(df, main, sub, family, work_type, search)

    if work_df.empty:
        st.warning("No BOQ items found for the current filters.")
        st.stop()

    display_df = work_df[["_row_id", "No", "Description", "Unit", "QTY", "U/R (SAR)", "T/P (SAR)"]].copy()

    st.markdown("### Selected BOQ Items")
    edited_df = st.data_editor(
        display_df,
        use_container_width=True,
        height=500,
        hide_index=True,
        disabled=["_row_id", "No", "Description", "Unit", "T/P (SAR)"],
        column_config={
            "_row_id": st.column_config.NumberColumn("ID", width="small"),
            "No": st.column_config.NumberColumn("No", width="small"),
            "Description": st.column_config.TextColumn("Description", width="large"),
            "Unit": st.column_config.TextColumn("Unit", width="small"),
            "QTY": st.column_config.NumberColumn("QTY", min_value=0.0, step=1.0),
            "U/R (SAR)": st.column_config.NumberColumn("U/R (SAR)", min_value=0.0, step=10.0, format="%.2f"),
            "T/P (SAR)": st.column_config.NumberColumn("T/P (SAR)", format="%.2f"),
        },
    )

    edited_df["T/P (SAR)"] = pd.to_numeric(edited_df["QTY"], errors="coerce").fillna(0) * pd.to_numeric(edited_df["U/R (SAR)"], errors="coerce").fillna(0)

    total = float(edited_df["T/P (SAR)"].sum())
    selection_text = f"{main} / {sub} / {family} / {work_type}"

    st.session_state.last_boq_df = edited_df.copy()
    st.session_state.last_selection_text = selection_text

    c1, c2, c3 = st.columns(3)
    with c1:
        card("Visible Items", f"{len(edited_df):,}", "Rows currently shown")
    with c2:
        card("Visible Total", money(total), "Total for current selection", gold=True)
    with c3:
        card("Selection", selection_text, "Current category path")

    st.markdown("### BOQ Actions")
    action_col1, action_col2 = st.columns([2, 1])

    with action_col1:
        item_options = [
            f"{int(row['No'])} | {row['Description'][:85]} | {row['Unit']} | SAR {float(row['U/R (SAR)']):,.2f}"
            for _, row in edited_df.iterrows()
        ]
        selected_option = st.selectbox("Select item for Rate Analysis", item_options)
        selected_index = item_options.index(selected_option)
        selected_row = edited_df.iloc[selected_index]

    with action_col2:
        st.write("")
        st.write("")
        if st.button("Send to Rate Analysis", type="primary"):
            st.session_state.rate_item = {
                "description": str(selected_row["Description"]),
                "unit": str(selected_row["Unit"]),
                "unit_rate": float(selected_row["U/R (SAR)"]),
                "no": str(selected_row["No"]),
            }
            st.success("Selected item sent to Rate Analysis page.")

    export_df = edited_df[["No", "Description", "Unit", "QTY", "U/R (SAR)", "T/P (SAR)"]].copy()

    excel_bytes = build_boq_export_excel(
        export_df=export_df,
        project_info=st.session_state.project_info,
        selection_text=selection_text,
    )

    st.download_button(
        label="⬇️ Export BOQ Excel",
        data=excel_bytes,
        file_name=f"KETKAT_BOQ_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

st.markdown('<div class="footer-note">KETKAT Hook-up & Installation Manager | Streamlit Web MVP V3</div>', unsafe_allow_html=True)